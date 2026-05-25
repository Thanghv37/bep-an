#report/views
from calendar import monthrange
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.db.models import Sum
from finance.models import DailyPurchase
from core.views import (
    get_registered_count,
    get_price_breakdown_for_date,
)
from accounts.permissions import can_view_report, can_export_report
def staff_required(user):
    return user.is_staff or user.is_superuser


def build_purchase_split_map(start_date, end_date):
    """Trả về (food_map, spice_map): chi phí theo ngày, tách theo phân loại hóa đơn
    (main → thực phẩm, extra → gia vị)."""
    rows = DailyPurchase.objects.filter(
        date__range=(start_date, end_date),
        status=DailyPurchase.STATUS_APPROVED
    ).values('date', 'purchase_type').annotate(
        total_cost=Sum('actual_cost')
    )

    food_map = {}
    spice_map = {}

    for row in rows:
        cost = int(row['total_cost'] or 0)
        if row['purchase_type'] == DailyPurchase.PURCHASE_TYPE_EXTRA:
            spice_map[row['date']] = spice_map.get(row['date'], 0) + cost
        else:
            food_map[row['date']] = food_map.get(row['date'], 0) + cost

    return food_map, spice_map


def build_purchase_map(start_date, end_date):
    """Tổng chi phí theo ngày (gộp cả 2 phân loại). Giữ để các hàm cũ dùng."""
    food_map, spice_map = build_purchase_split_map(start_date, end_date)
    combined = {}
    for d, v in food_map.items():
        combined[d] = combined.get(d, 0) + v
    for d, v in spice_map.items():
        combined[d] = combined.get(d, 0) + v
    return combined


@login_required
@user_passes_test(can_view_report)
def report_dashboard(request):
    today = date.today()

    form_type = request.GET.get('form_type', 'main')

    # =========================
    # FORM 1: BÁO CÁO TỔNG HỢP
    # =========================
    if form_type == 'main':
        view_type = request.GET.get('view_type', 'month')

        try:
            selected_month = int(request.GET.get('month', today.month))
        except (ValueError, TypeError):
            selected_month = today.month

        try:
            selected_year = int(request.GET.get('year', today.year))
        except (ValueError, TypeError):
            selected_year = today.year

        if selected_month < 1 or selected_month > 12:
            selected_month = today.month

        try:
            selected_date_str = request.GET.get('selected_date')
            selected_date = date.fromisoformat(selected_date_str) if selected_date_str else today
        except ValueError:
            selected_date = today

        balance_chart_type = request.GET.get('balance_chart_type', 'daily_in_month')

    # =========================
    # FORM 2: BIỂU ĐỒ CHÊNH LỆCH
    # =========================
    else:
        view_type = request.GET.get('view_type_state', 'month')

        try:
            selected_month = int(request.GET.get('month_state', today.month))
        except (ValueError, TypeError):
            selected_month = today.month

        try:
            selected_year = int(request.GET.get('year_state', today.year))
        except (ValueError, TypeError):
            selected_year = today.year

        if selected_month < 1 or selected_month > 12:
            selected_month = today.month

        try:
            selected_date_str = request.GET.get('selected_date_state')
            selected_date = date.fromisoformat(selected_date_str) if selected_date_str else today
        except ValueError:
            selected_date = today

        balance_chart_type = request.GET.get('balance_chart_type', 'daily_in_month')

    # =========================
    # XÁC ĐỊNH KHOẢNG THỜI GIAN CHÍNH
    # =========================
    if view_type == 'week':
        start_date = selected_date - timedelta(days=selected_date.weekday())
        end_date = start_date + timedelta(days=6)
    else:
        start_date = date(selected_year, selected_month, 1)
        end_day = monthrange(selected_year, selected_month)[1]
        end_date = date(selected_year, selected_month, end_day)

    purchase_food_map, purchase_spice_map = build_purchase_split_map(start_date, end_date)
    purchase_detail_map = {}

    all_purchases = DailyPurchase.objects.filter(
        date__range=(start_date, end_date),
        status=DailyPurchase.STATUS_APPROVED
    ).select_related('created_by')

    for p in all_purchases:
        purchase_detail_map.setdefault(p.date, []).append(p)

    daily_rows = []
    chart_labels = []
    income_food_data = []
    income_spice_data = []
    expense_food_data = []
    expense_spice_data = []
    balance_food_data = []
    balance_spice_data = []

    total_income_food = 0
    total_income_spice = 0
    total_expense_food = 0
    total_expense_spice = 0
    total_balance_food = 0
    total_balance_spice = 0

    current_date = start_date
    while current_date <= end_date:
        registered_count = get_registered_count(current_date)
        breakdown = get_price_breakdown_for_date(current_date)
        food_cost = purchase_food_map.get(current_date, 0)
        spice_cost = purchase_spice_map.get(current_date, 0)

        if breakdown:
            meal_price = breakdown['meal']
            food_price = breakdown['food']
            spice_price = breakdown['spice']
            income_food = registered_count * food_price
            income_spice = registered_count * spice_price
            income = income_food + income_spice
            balance_food = income_food - food_cost
            balance_spice = income_spice - spice_cost
        else:
            meal_price = food_price = spice_price = None
            income_food = income_spice = income = None
            balance_food = balance_spice = None

        expense = food_cost + spice_cost
        balance = (income - expense) if income is not None else None

        row = {
            'date': current_date,
            'registered_count': registered_count,
            'meal_price': meal_price,
            'food_price': food_price,
            'spice_price': spice_price,
            'income': income,
            'income_food': income_food,
            'income_spice': income_spice,
            'expense': expense,
            'expense_food': food_cost,
            'expense_spice': spice_cost,
            'balance': balance,
            'balance_food': balance_food,
            'balance_spice': balance_spice,
            'purchase_list': purchase_detail_map.get(current_date, []),
        }
        daily_rows.append(row)

        chart_labels.append(current_date.strftime('%d/%m'))
        income_food_data.append(income_food)
        income_spice_data.append(income_spice)
        expense_food_data.append(food_cost)
        expense_spice_data.append(spice_cost)
        balance_food_data.append(balance_food)
        balance_spice_data.append(balance_spice)

        if income_food is not None:
            total_income_food += income_food
            total_income_spice += income_spice

        total_expense_food += food_cost
        total_expense_spice += spice_cost

        if balance_food is not None:
            total_balance_food += balance_food
            total_balance_spice += balance_spice

        current_date += timedelta(days=1)

    total_income = total_income_food + total_income_spice
    total_expense = total_expense_food + total_expense_spice
    total_balance = total_balance_food + total_balance_spice

    # =========================
    # BIỂU ĐỒ CHÊNH LỆCH RIÊNG — tách theo Thực phẩm & Gia vị
    # =========================
    balance_chart_labels = []
    balance_food_chart_values = []
    balance_spice_chart_values = []

    if balance_chart_type == 'monthly_in_year':
        year_start = date(selected_year, 1, 1)
        year_end = date(selected_year, 12, 31)
        year_food_map, year_spice_map = build_purchase_split_map(year_start, year_end)

        for month in range(1, 13):
            month_start = date(selected_year, month, 1)
            month_end = date(selected_year, month, monthrange(selected_year, month)[1])

            month_balance_food = 0
            month_balance_spice = 0
            has_any = False

            d = month_start
            while d <= month_end:
                cnt = get_registered_count(d)
                bd = get_price_breakdown_for_date(d)
                day_food_cost = year_food_map.get(d, 0)
                day_spice_cost = year_spice_map.get(d, 0)

                if bd:
                    month_balance_food += cnt * bd['food'] - day_food_cost
                    month_balance_spice += cnt * bd['spice'] - day_spice_cost
                    has_any = True

                d += timedelta(days=1)

            balance_chart_labels.append(f'T{month}')
            balance_food_chart_values.append(month_balance_food if has_any else None)
            balance_spice_chart_values.append(month_balance_spice if has_any else None)
    elif balance_chart_type == 'daily_in_week':
        # X-axis = 5 ngày làm việc T2 → T6 của tuần chứa selected_date (bỏ T7 & CN).
        week_start = selected_date - timedelta(days=selected_date.weekday())
        week_end = week_start + timedelta(days=4)
        week_food_map, week_spice_map = build_purchase_split_map(week_start, week_end)

        d = week_start
        while d <= week_end:
            cnt = get_registered_count(d)
            bd = get_price_breakdown_for_date(d)
            day_food_cost = week_food_map.get(d, 0)
            day_spice_cost = week_spice_map.get(d, 0)

            if bd:
                bf = cnt * bd['food'] - day_food_cost
                bs = cnt * bd['spice'] - day_spice_cost
            else:
                bf = bs = None

            balance_chart_labels.append(d.strftime('%d/%m'))
            balance_food_chart_values.append(bf)
            balance_spice_chart_values.append(bs)
            d += timedelta(days=1)
    else:
        month_start = date(selected_year, selected_month, 1)
        month_end = date(selected_year, selected_month, monthrange(selected_year, selected_month)[1])
        month_food_map, month_spice_map = build_purchase_split_map(month_start, month_end)

        d = month_start
        while d <= month_end:
            cnt = get_registered_count(d)
            bd = get_price_breakdown_for_date(d)
            day_food_cost = month_food_map.get(d, 0)
            day_spice_cost = month_spice_map.get(d, 0)

            if bd:
                bf = cnt * bd['food'] - day_food_cost
                bs = cnt * bd['spice'] - day_spice_cost
            else:
                bf = bs = None

            balance_chart_labels.append(d.strftime('%d/%m'))
            balance_food_chart_values.append(bf)
            balance_spice_chart_values.append(bs)
            d += timedelta(days=1)

    # Lũy kế chênh lệch (cộng dồn từ đầu kỳ tới điểm hiện tại).
    # None coi như 0 — không break đường lũy kế khi vài ngày chưa cấu hình giá.
    cumulative_food_chart_values = []
    cumulative_spice_chart_values = []
    cumulative_total_chart_values = []
    acc_food = 0
    acc_spice = 0
    for bf, bs in zip(balance_food_chart_values, balance_spice_chart_values):
        if bf is not None:
            acc_food += bf
        if bs is not None:
            acc_spice += bs
        cumulative_food_chart_values.append(acc_food)
        cumulative_spice_chart_values.append(acc_spice)
        cumulative_total_chart_values.append(acc_food + acc_spice)

    # Che các điểm tương lai (chưa tới) — tránh hiển thị 0 vô nghĩa.
    if balance_chart_type == 'monthly_in_year':
        # X-axis = 12 tháng; index 0=T1 ... index 11=T12. Giữ index < today.month nếu selected_year == today.year.
        if selected_year > today.year:
            mask_from = 0
        elif selected_year < today.year:
            mask_from = 12
        else:
            mask_from = today.month
    elif balance_chart_type == 'daily_in_week':
        # X-axis = 5 ngày làm việc T2..T6. Mask các ngày > today.
        week_start = selected_date - timedelta(days=selected_date.weekday())
        if week_start > today:
            mask_from = 0
        elif week_start + timedelta(days=4) < today:
            mask_from = 5
        else:
            # today nằm trong khoảng T2..T6 (hoặc cuối tuần — clamp về 5).
            mask_from = min((today - week_start).days + 1, 5)
    else:
        # daily_in_month — X-axis = các ngày trong selected_month/selected_year.
        if selected_year > today.year or (selected_year == today.year and selected_month > today.month):
            mask_from = 0
        elif selected_year < today.year or (selected_year == today.year and selected_month < today.month):
            mask_from = len(balance_chart_labels)
        else:
            mask_from = today.day

    for i in range(mask_from, len(balance_chart_labels)):
        balance_food_chart_values[i] = None
        balance_spice_chart_values[i] = None
        cumulative_food_chart_values[i] = None
        cumulative_spice_chart_values[i] = None
        cumulative_total_chart_values[i] = None

    month_choices = list(range(1, 13))
    year_choices = list(range(today.year - 2, today.year + 3))

    context = {
        'form_type': form_type,
        'view_type': view_type,
        'balance_chart_type': balance_chart_type,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_date': selected_date.isoformat() if selected_date else '',
        'month_choices': month_choices,
        'year_choices': year_choices,
        'daily_rows': daily_rows,
        'total_income': total_income,
        'total_income_food': total_income_food,
        'total_income_spice': total_income_spice,
        'total_expense': total_expense,
        'total_expense_food': total_expense_food,
        'total_expense_spice': total_expense_spice,
        'total_balance': total_balance,
        'total_balance_food': total_balance_food,
        'total_balance_spice': total_balance_spice,
        'chart_labels': chart_labels,
        'income_food_data': income_food_data,
        'income_spice_data': income_spice_data,
        'expense_food_data': expense_food_data,
        'expense_spice_data': expense_spice_data,
        'balance_food_data': balance_food_data,
        'balance_spice_data': balance_spice_data,
        'balance_chart_labels': balance_chart_labels,
        'balance_food_chart_values': balance_food_chart_values,
        'balance_spice_chart_values': balance_spice_chart_values,
        'cumulative_food_chart_values': cumulative_food_chart_values,
        'cumulative_spice_chart_values': cumulative_spice_chart_values,
        'cumulative_total_chart_values': cumulative_total_chart_values,
        'start_date': start_date,
        'end_date': end_date,
        'can_export_report': can_export_report(request.user),
    }
    return render(request, 'reports/report_dashboard.html', context)


# ================================================================
# EXPORT VIEWS
# ================================================================
import io
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from finance.models import PurchaseExtraItem
from reviews.models import DishReview, MealReview
from django.db.models import Avg, Count, Q


# ---------- helpers ----------
HEADER_FILL  = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL   = PatternFill("solid", fgColor="DBEAFE")
TOTAL_FONT   = Font(bold=True, size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="EFF6FF")
CENTER       = Alignment(horizontal="center", vertical="center")
WRAP         = Alignment(wrap_text=True, vertical="top")

thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_vnd(val):
    if val is None: return ""
    return f"{int(val):,}".replace(",", ".")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _header_row(ws, cols, row=1):
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _get_period(request, prefix=""):
    """Parse period params; returns (period, start_date, end_date, label)."""
    today = date.today()
    period = request.GET.get(f"{prefix}period", "month")
    try: sel_date = date.fromisoformat(request.GET.get(f"{prefix}date", today.isoformat()))
    except: sel_date = today
    try: sel_month = int(request.GET.get(f"{prefix}month", today.month))
    except: sel_month = today.month
    try: sel_year  = int(request.GET.get(f"{prefix}year",  today.year))
    except: sel_year  = today.year

    if period == "week":
        start = sel_date - timedelta(days=sel_date.weekday())
        end   = start + timedelta(days=6)
        label = f"tuan_{start.strftime('%d%m%Y')}_{end.strftime('%d%m%Y')}"
    elif period == "month":
        start = date(sel_year, sel_month, 1)
        end   = date(sel_year, sel_month, monthrange(sel_year, sel_month)[1])
        label = f"thang_{sel_month:02d}_{sel_year}"
    elif period == "year":
        start = date(sel_year, 1, 1)
        end   = date(sel_year, 12, 31)
        label = f"nam_{sel_year}"
    else:  # day
        start = end = sel_date
        label = f"ngay_{sel_date.strftime('%d%m%Y')}"

    return period, start, end, label, sel_year, sel_month


# ================================================================
# 1. BÁO CÁO THU – CHI
# ================================================================
@login_required
@user_passes_test(can_export_report)
def export_revenue_report(request):
    period, start_date, end_date, label, sel_year, sel_month = _get_period(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thu - Chi"

    title_text = f"BÁO CÁO THU – CHI  ({start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')})"

    if period == "year":
        # Xuất theo tháng: 8 cột tách Thực phẩm / Gia vị
        ws.merge_cells("A1:H1")
        t = ws["A1"]
        t.value = title_text
        t.font = Font(bold=True, size=13, color="1E40AF")
        t.alignment = CENTER
        ws.row_dimensions[1].height = 28

        cols = [
            "Tháng", "Tổng suất ĐK",
            "Thu Thực phẩm (VNĐ)", "Thu Gia vị (VNĐ)",
            "Chi Thực phẩm (VNĐ)", "Chi Gia vị (VNĐ)",
            "Chênh Thực phẩm (VNĐ)", "Chênh Gia vị (VNĐ)",
        ]
        _header_row(ws, cols, row=2)
        ws.row_dimensions[2].height = 32

        grand_in_food = grand_in_spice = 0
        grand_ex_food = grand_ex_spice = 0
        grand_bal_food = grand_bal_spice = 0
        row_idx = 3

        for m in range(1, 13):
            ms = date(sel_year, m, 1)
            me = date(sel_year, m, monthrange(sel_year, m)[1])

            m_in_food = m_in_spice = 0
            m_bal_food = m_bal_spice = 0
            m_people = 0

            food_map_m, spice_map_m = build_purchase_split_map(ms, me)

            d = ms
            while d <= me:
                cnt = get_registered_count(d)
                bd = get_price_breakdown_for_date(d)
                day_fc = food_map_m.get(d, 0)
                day_sc = spice_map_m.get(d, 0)

                if bd and cnt:
                    m_in_food += cnt * bd['food']
                    m_in_spice += cnt * bd['spice']
                    m_bal_food += cnt * bd['food'] - day_fc
                    m_bal_spice += cnt * bd['spice'] - day_sc
                    m_people += cnt

                d += timedelta(days=1)

            m_ex_food = sum(food_map_m.values())
            m_ex_spice = sum(spice_map_m.values())

            grand_in_food += m_in_food
            grand_in_spice += m_in_spice
            grand_ex_food += m_ex_food
            grand_ex_spice += m_ex_spice
            grand_bal_food += m_bal_food
            grand_bal_spice += m_bal_spice

            row_data = [
                f"Tháng {m}/{sel_year}", m_people,
                _fmt_vnd(m_in_food), _fmt_vnd(m_in_spice),
                _fmt_vnd(m_ex_food), _fmt_vnd(m_ex_spice),
                _fmt_vnd(m_bal_food), _fmt_vnd(m_bal_spice),
            ]
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = BORDER
                cell.alignment = CENTER if c != 1 else Alignment(horizontal="left", vertical="center")
                if c == 7 and m_bal_food < 0:
                    cell.font = Font(color="DC2626")
                if c == 8 and m_bal_spice < 0:
                    cell.font = Font(color="DC2626")
            row_idx += 1

        for c, v in enumerate([
            "TỔNG KẾT", "",
            _fmt_vnd(grand_in_food), _fmt_vnd(grand_in_spice),
            _fmt_vnd(grand_ex_food), _fmt_vnd(grand_ex_spice),
            _fmt_vnd(grand_bal_food), _fmt_vnd(grand_bal_spice),
        ], 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Ô cuối: tổng chênh lệch = Chênh TP + Chênh GV (đặt sau cột Chênh GV)
        total_diff = grand_bal_food + grand_bal_spice
        sum_cell = ws.cell(row=row_idx, column=9, value=f"Tổng chênh lệch: {_fmt_vnd(total_diff)}")
        sum_cell.fill = PatternFill("solid", fgColor="1E40AF")
        sum_cell.font = Font(bold=True, size=11, color="FFFFFF")
        sum_cell.border = BORDER
        sum_cell.alignment = CENTER

    else:
        # Xuất theo ngày (tuần/tháng): 11 cột tách Thực phẩm / Gia vị
        ws.merge_cells("A1:K1")
        t = ws["A1"]
        t.value = title_text
        t.font = Font(bold=True, size=13, color="1E40AF")
        t.alignment = CENTER
        ws.row_dimensions[1].height = 28

        cols = [
            "Ngày", "Thứ", "Số người ĐK",
            "Giá TP (VNĐ)", "Giá GV (VNĐ)",
            "Thu TP (VNĐ)", "Thu GV (VNĐ)",
            "Chi TP (VNĐ)", "Chi GV (VNĐ)",
            "Chênh TP (VNĐ)", "Chênh GV (VNĐ)",
        ]
        _header_row(ws, cols, row=2)
        ws.row_dimensions[2].height = 32

        WEEKDAYS = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
        food_map_p, spice_map_p = build_purchase_split_map(start_date, end_date)

        grand_in_food = grand_in_spice = 0
        grand_ex_food = grand_ex_spice = 0
        grand_bal_food = grand_bal_spice = 0
        row_idx = 3
        d = start_date
        while d <= end_date:
            cnt = get_registered_count(d)
            bd = get_price_breakdown_for_date(d)
            food_cost = food_map_p.get(d, 0)
            spice_cost = spice_map_p.get(d, 0)

            if bd:
                food_price = bd['food']
                spice_price = bd['spice']
                in_food = cnt * food_price
                in_spice = cnt * spice_price
                bal_food = in_food - food_cost
                bal_spice = in_spice - spice_cost
            else:
                food_price = spice_price = None
                in_food = in_spice = None
                bal_food = bal_spice = None

            if in_food is not None:
                grand_in_food += in_food
                grand_in_spice += in_spice
            grand_ex_food += food_cost
            grand_ex_spice += spice_cost
            if bal_food is not None:
                grand_bal_food += bal_food
                grand_bal_spice += bal_spice

            row_data = [
                d.strftime("%d/%m/%Y"),
                WEEKDAYS[d.weekday()],
                cnt or "",
                _fmt_vnd(food_price), _fmt_vnd(spice_price),
                _fmt_vnd(in_food), _fmt_vnd(in_spice),
                _fmt_vnd(food_cost), _fmt_vnd(spice_cost),
                _fmt_vnd(bal_food), _fmt_vnd(bal_spice),
            ]
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = BORDER
                cell.alignment = CENTER if c > 1 else Alignment(horizontal="left", vertical="center")
                if c == 10 and bal_food is not None and bal_food < 0:
                    cell.font = Font(color="DC2626")
                if c == 11 and bal_spice is not None and bal_spice < 0:
                    cell.font = Font(color="DC2626")
            row_idx += 1
            d += timedelta(days=1)

        for c, v in enumerate([
            "TỔNG KẾT", "", "",
            "", "",
            _fmt_vnd(grand_in_food), _fmt_vnd(grand_in_spice),
            _fmt_vnd(grand_ex_food), _fmt_vnd(grand_ex_spice),
            _fmt_vnd(grand_bal_food), _fmt_vnd(grand_bal_spice),
        ], 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Ô cuối: tổng chênh lệch = Chênh TP + Chênh GV (đặt sau cột Chênh GV)
        total_diff = grand_bal_food + grand_bal_spice
        sum_cell = ws.cell(row=row_idx, column=12, value=f"Tổng chênh lệch: {_fmt_vnd(total_diff)}")
        sum_cell.fill = PatternFill("solid", fgColor="1E40AF")
        sum_cell.font = Font(bold=True, size=11, color="FFFFFF")
        sum_cell.border = BORDER
        sum_cell.alignment = CENTER

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="bao_cao_thu_chi_{label}.xlsx"'
    return response


# ================================================================
# 2. BÁO CÁO CHI PHÍ
# ================================================================
@login_required
@user_passes_test(can_export_report)
def export_cost_report(request):
    period, start_date, end_date, label, sel_year, sel_month = _get_period(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi phí"

    title_text = f"BÁO CÁO CHI PHÍ  ({start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')})"
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = title_text
    t.font = Font(bold=True, size=13, color="1E40AF")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    if period in ("day", "week"):
        # Chi tiết từng mặt hàng
        cols = ["Ngày", "Loại mua", "Tên mặt hàng", "Số lượng", "Đơn vị", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)"]
        _header_row(ws, cols, row=2)

        items = PurchaseExtraItem.objects.filter(
            date__range=(start_date, end_date),
            purchase__status=DailyPurchase.STATUS_APPROVED
        ).select_related("purchase").order_by("date", "purchase_id")

        row_idx = 3
        grand_total = 0
        for item in items:
            total_line = float(item.quantity or 0) * float(item.unit_price or 0)
            grand_total += total_line
            row_data = [
                item.date.strftime("%d/%m/%Y"),
                item.purchase.get_purchase_type_display(),
                item.ingredient_name,
                float(item.quantity or 0),
                item.unit or "",
                _fmt_vnd(item.unit_price),
                _fmt_vnd(total_line),
            ]
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = BORDER
                cell.alignment = CENTER if c in (4, 5, 6, 7) else Alignment(horizontal="left", vertical="center")
            row_idx += 1

        # Phiếu không có chi tiết (chưa quét AI)
        purchases_no_items = DailyPurchase.objects.filter(
            date__range=(start_date, end_date),
            status=DailyPurchase.STATUS_APPROVED,
            extra_items__isnull=True
        ).distinct()
        for p in purchases_no_items:
            row_data = [p.date.strftime("%d/%m/%Y"), p.get_purchase_type_display(), "(Chưa quét AI – chỉ có tổng)", "", "", "", _fmt_vnd(p.actual_cost)]
            grand_total += float(p.actual_cost or 0)
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = BORDER
                cell.font = Font(italic=True, color="6B7280")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            row_idx += 1

        # Tổng
        for c, v in enumerate(["TỔNG CHI PHÍ", "", "", "", "", "", _fmt_vnd(grand_total)], 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER

    else:
        # Tháng/Năm: tổng hợp, tách theo Thực phẩm / Gia vị
        def _agg_by_type(start, end):
            """Trả về (cnt_tp, total_tp, cnt_gv, total_gv) trong khoảng [start, end]."""
            rows = DailyPurchase.objects.filter(
                date__range=(start, end), status=DailyPurchase.STATUS_APPROVED
            ).values('purchase_type').annotate(total=Sum('actual_cost'), cnt=Count('id'))
            cnt_tp = cnt_gv = 0
            tot_tp = tot_gv = 0
            for r in rows:
                if r['purchase_type'] == DailyPurchase.PURCHASE_TYPE_EXTRA:
                    cnt_gv += r['cnt'] or 0
                    tot_gv += int(r['total'] or 0)
                else:
                    cnt_tp += r['cnt'] or 0
                    tot_tp += int(r['total'] or 0)
            return cnt_tp, tot_tp, cnt_gv, tot_gv

        if period == "year":
            cols = [
                "Tháng",
                "TP - Số phiếu", "TP - Tổng chi (VNĐ)",
                "GV - Số phiếu", "GV - Tổng chi (VNĐ)",
                "Tổng phiếu", "Tổng chi (VNĐ)",
            ]
            _header_row(ws, cols, row=2)
            ws.row_dimensions[2].height = 32
            row_idx = 3
            g_cnt_tp = g_tot_tp = g_cnt_gv = g_tot_gv = 0
            for m in range(1, 13):
                ms = date(sel_year, m, 1)
                me = date(sel_year, m, monthrange(sel_year, m)[1])
                cnt_tp, tot_tp, cnt_gv, tot_gv = _agg_by_type(ms, me)
                g_cnt_tp += cnt_tp; g_tot_tp += tot_tp
                g_cnt_gv += cnt_gv; g_tot_gv += tot_gv
                row_data = [
                    f"Tháng {m}/{sel_year}",
                    cnt_tp, _fmt_vnd(tot_tp),
                    cnt_gv, _fmt_vnd(tot_gv),
                    cnt_tp + cnt_gv, _fmt_vnd(tot_tp + tot_gv),
                ]
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=c, value=v)
                    cell.border = BORDER
                    cell.alignment = CENTER
                row_idx += 1
            for c, v in enumerate([
                "TỔNG KẾT",
                g_cnt_tp, _fmt_vnd(g_tot_tp),
                g_cnt_gv, _fmt_vnd(g_tot_gv),
                g_cnt_tp + g_cnt_gv, _fmt_vnd(g_tot_tp + g_tot_gv),
            ], 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER; cell.alignment = CENTER
        else:
            # month
            cols = [
                "Ngày",
                "TP - Số phiếu", "TP - Tổng chi (VNĐ)",
                "GV - Số phiếu", "GV - Tổng chi (VNĐ)",
                "Tổng chi (VNĐ)",
            ]
            _header_row(ws, cols, row=2)
            ws.row_dimensions[2].height = 32
            row_idx = 3
            g_cnt_tp = g_tot_tp = g_cnt_gv = g_tot_gv = 0
            d = start_date
            while d <= end_date:
                cnt_tp, tot_tp, cnt_gv, tot_gv = _agg_by_type(d, d)
                if cnt_tp or cnt_gv:
                    g_cnt_tp += cnt_tp; g_tot_tp += tot_tp
                    g_cnt_gv += cnt_gv; g_tot_gv += tot_gv
                    row_data = [
                        d.strftime("%d/%m/%Y"),
                        cnt_tp, _fmt_vnd(tot_tp),
                        cnt_gv, _fmt_vnd(tot_gv),
                        _fmt_vnd(tot_tp + tot_gv),
                    ]
                    for c, v in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=c, value=v)
                        cell.border = BORDER; cell.alignment = CENTER
                    row_idx += 1
                d += timedelta(days=1)
            for c, v in enumerate([
                "TỔNG KẾT",
                g_cnt_tp, _fmt_vnd(g_tot_tp),
                g_cnt_gv, _fmt_vnd(g_tot_gv),
                _fmt_vnd(g_tot_tp + g_tot_gv),
            ], 1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER; cell.alignment = CENTER

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="bao_cao_chi_phi_{label}.xlsx"'
    return response


# ================================================================
# 3. BÁO CÁO ĐÁNH GIÁ MÓN ĂN
# ================================================================
@login_required
@user_passes_test(can_export_report)
def export_review_report(request):
    period, start_date, end_date, label, sel_year, sel_month = _get_period(request)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Thống kê Like/Dislike ----
    ws1 = wb.active
    ws1.title = "Thống kê đánh giá"

    title_text = f"BÁO CÁO ĐÁNH GIÁ MÓN ĂN  ({start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')})"
    ws1.merge_cells("A1:E1")
    t = ws1["A1"]
    t.value = title_text
    t.font = Font(bold=True, size=13, color="1E40AF")
    t.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    cols = ["Tên món ăn", "Điểm TB (⭐)", "Tổng lượt đánh giá", "Số 5⭐", "Số 1⭐"]
    _header_row(ws1, cols, row=2)

    dish_stats = DishReview.objects.filter(
        meal_review__date__range=(start_date, end_date)
    ).values("dish__name").annotate(
        avg_rating = Avg("rating"),
        total      = Count("id"),
        top_count  = Count("id", filter=Q(rating=5)),
        low_count  = Count("id", filter=Q(rating=1)),
    ).order_by("-avg_rating")

    row_idx = 3
    for stat in dish_stats:
        avg = round(stat['avg_rating'], 2) if stat['avg_rating'] is not None else 0
        row_data = [stat['dish__name'], avg, stat['total'], stat['top_count'], stat['low_count']]
        for c, v in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else Alignment(horizontal="left", vertical="center")
            if c == 2:
                if avg >= 4:
                    cell.font = Font(color="16A34A", bold=True)
                elif avg and avg < 3:
                    cell.font = Font(color="DC2626", bold=True)
        row_idx += 1

    _auto_width(ws1)

    # ---- Sheet 2: Chi tiết góp ý ----
    ws2 = wb.create_sheet("Chi tiết góp ý")
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value = "CHI TIẾT GÓP Ý TYPING"
    t2.font = Font(bold=True, size=13, color="1E40AF")
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 28

    _header_row(ws2, ["Ngày", "Người đánh giá", "Nội dung góp ý"], row=2)

    comments = MealReview.objects.filter(
        date__range=(start_date, end_date)
    ).exclude(comment="").select_related("user").order_by("-date")

    row_idx = 3
    for review in comments:
        user_display = review.user.username if review.user else "Ẩn danh"
        row_data = [review.date.strftime("%d/%m/%Y"), user_display, review.comment]
        for c, v in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP if c == 3 else Alignment(horizontal="left", vertical="top")
        ws2.row_dimensions[row_idx].height = max(30, min(len(review.comment) // 3, 120))
        row_idx += 1

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="bao_cao_danh_gia_{label}.xlsx"'
    return response