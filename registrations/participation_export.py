"""Tạo file Excel báo cáo Tham gia + gửi qua NetChat DM.

Tách riêng khỏi views.py để dùng chung giữa endpoint download và endpoint
gửi NetChat (và sau này nếu có scheduled job).
"""

import io
import json
import re

import requests

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from accounts.models import UserProfile
from core.models import SystemConfig


KEY_RECIPIENTS = 'participation_export_recipients'
KEY_SEND_TIME = 'participation_export_send_time'
KEY_SEND_MODE = 'participation_export_send_mode'      # 'dm' | 'channel'
KEY_CHANNEL_ID = 'participation_export_channel_id'
KEY_SEND_DAYS = 'participation_export_send_days'      # JSON list of int 0-6
# Mặc định gửi T2-T6 (weekday 0..4), không gửi T7/CN.
DEFAULT_SEND_DAYS = [0, 1, 2, 3, 4]


# ---------- Config trong SystemConfig ----------

def get_recipients():
    cfg = SystemConfig.objects.filter(key=KEY_RECIPIENTS).first()
    if not cfg or not cfg.value:
        return []
    try:
        data = json.loads(cfg.value)
        if isinstance(data, list):
            return [str(c).strip() for c in data if str(c).strip()]
    except (ValueError, TypeError):
        pass
    return []


def set_recipients(codes):
    cleaned = []
    seen = set()
    for c in codes:
        c = (c or '').strip()
        if c and c not in seen:
            seen.add(c)
            cleaned.append(c)
    SystemConfig.objects.update_or_create(
        key=KEY_RECIPIENTS,
        defaults={'value': json.dumps(cleaned, ensure_ascii=False)},
    )
    return cleaned


def get_send_time():
    cfg = SystemConfig.objects.filter(key=KEY_SEND_TIME).first()
    val = (cfg.value if cfg else '') or ''
    # Chỉ chấp nhận HH:MM, mặc định 13:00.
    if re.match(r'^\d{2}:\d{2}$', val):
        return val
    return '13:00'


def set_send_time(value):
    val = (value or '').strip()
    if not re.match(r'^\d{2}:\d{2}$', val):
        raise ValueError('Thời gian phải có dạng HH:MM (24h).')
    SystemConfig.objects.update_or_create(
        key=KEY_SEND_TIME,
        defaults={'value': val},
    )
    return val


def get_send_mode():
    """Hình thức gửi báo cáo: 'dm' (gửi từng người - mặc định) hoặc 'channel'."""
    cfg = SystemConfig.objects.filter(key=KEY_SEND_MODE).first()
    val = (cfg.value if cfg else '') or ''
    return 'channel' if val == 'channel' else 'dm'


def set_send_mode(value):
    val = 'channel' if (value or '').strip() == 'channel' else 'dm'
    SystemConfig.objects.update_or_create(
        key=KEY_SEND_MODE,
        defaults={'value': val},
    )
    return val


def get_send_days():
    """Trả về list weekday gửi báo cáo (0=Thứ 2 ... 6=CN). Mặc định T2-T6."""
    cfg = SystemConfig.objects.filter(key=KEY_SEND_DAYS).first()
    if not cfg or not cfg.value:
        return list(DEFAULT_SEND_DAYS)
    try:
        data = json.loads(cfg.value)
        if isinstance(data, list):
            cleaned = sorted({int(d) for d in data if 0 <= int(d) <= 6})
            return cleaned if cleaned else list(DEFAULT_SEND_DAYS)
    except (ValueError, TypeError):
        pass
    return list(DEFAULT_SEND_DAYS)


def set_send_days(values):
    """Lưu list weekday. `values` có thể là list int / list str / CSV string."""
    if isinstance(values, str):
        values = [v.strip() for v in values.split(',') if v.strip()]
    cleaned = sorted({int(v) for v in (values or []) if str(v).strip().isdigit() and 0 <= int(v) <= 6})
    SystemConfig.objects.update_or_create(
        key=KEY_SEND_DAYS,
        defaults={'value': json.dumps(cleaned)},
    )
    return cleaned


def get_channel_id():
    cfg = SystemConfig.objects.filter(key=KEY_CHANNEL_ID).first()
    return (cfg.value.strip() if cfg and cfg.value else '')


def set_channel_id(value):
    val = (value or '').strip()
    SystemConfig.objects.update_or_create(
        key=KEY_CHANNEL_ID,
        defaults={'value': val},
    )
    return val


# ---------- Thống kê & caption ----------

def count_statuses(rows):
    """Đếm số dòng theo từng trạng thái tham gia."""
    return {
        'valid': sum(1 for r in rows if r['status'] == 'valid'),
        'supplementary': sum(1 for r in rows if r['status'] == 'supplementary'),
        'not_attended': sum(1 for r in rows if r['status'] == 'not_attended'),
        'not_registered': sum(1 for r in rows if r['status'] == 'not_registered'),
        'no_profile': sum(1 for r in rows if r['status'] == 'no_profile'),
    }


def build_report_caption(target_date, rows=None):
    """Caption tin nhắn NetChat. Nếu có `rows` thì kèm số liệu tổng hợp +
    liệt kê tên người ở các nhóm cần chú ý (chưa điểm danh / chưa đăng ký /
    chưa có hồ sơ) để người nhận nắm ngay không cần mở file."""
    caption = f'📊 Báo cáo Tham gia ngày {target_date.strftime("%d-%m-%Y")}'
    if rows is None:
        return caption
    c = count_statuses(rows)

    def _names(status):
        return [r for r in rows if r['status'] == status]

    def _fmt(items):
        return ', '.join(
            f'{r["display_name"]} ({r["employee_code"]})' for r in items
        )

    total_qty = sum(r.get('quantity', 0) for r in rows)
    caption += f'\n- Tổng suất đăng ký: {total_qty}'

    caption += f'\n- Đã điểm danh: {c["valid"]}'

    supplementary = _names('supplementary')
    caption += f'\n- Đăng ký bổ sung: {len(supplementary)}'
    if supplementary:
        caption += f'\n   👉 {_fmt(supplementary)}'

    not_attended = _names('not_attended')
    caption += f'\n- Chưa điểm danh: {len(not_attended)}'
    if not_attended:
        caption += f'\n   👉 {_fmt(not_attended)}'

    not_registered = _names('not_registered')
    caption += f'\n- Chưa đăng ký: {len(not_registered)}'
    if not_registered:
        caption += f'\n   👉 {_fmt(not_registered)}'

    no_profile = _names('no_profile')
    if no_profile:
        caption += f'\n- Chưa có hồ sơ (đơn vị khác): {len(no_profile)}'
        caption += f'\n   👉 {_fmt(no_profile)}'

    return caption


# ---------- Build Excel ----------

def build_excel_bytes(target_date, rows):
    """Tạo file Excel binary từ rows. `rows` đến từ `_build_participation_rows`."""
    status_order = {'valid': 0, 'supplementary': 1, 'not_attended': 2, 'not_registered': 3, 'no_profile': 4}
    rows = sorted(rows, key=lambda r: (
        status_order.get(r['status'], 99),
        r['scan_time'] or 0,
        r['display_name'],
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tham gia'

    headers = ['STT', 'Họ và tên', 'Mã NV', 'Số suất ĐK', 'Đơn vị', 'Phòng ban', 'Thời gian quét', 'Trạng thái', 'Loại']
    n_cols = len(headers)
    last_col_letter = chr(ord('A') + n_cols - 1)

    # Row 1: Title
    title = f'DANH SÁCH THAM GIA NGÀY {target_date.strftime("%d-%m-%Y")}'
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Summary (đưa lên trên cùng theo yêu cầu — người xem nắm số liệu trước khi nhìn list)
    counts = count_statuses(rows)
    total_quantity = sum(r.get('quantity', 0) for r in rows)
    summary_cells = [
        ('Tổng:', True),
        (f"Đã điểm danh: {counts['valid']}", False),
        (f"Đăng ký bổ sung: {counts['supplementary']}", False),
        (f"Chưa điểm danh: {counts['not_attended']}", False),
        (f"Chưa đăng ký: {counts['not_registered']}", False),
        (f"Chưa có hồ sơ: {counts.get('no_profile', 0)}", False),
        (f"Tổng suất ĐK: {total_quantity}", True),
    ]
    for col_idx, (text, bold) in enumerate(summary_cells, start=1):
        cell = ws.cell(row=2, column=col_idx, value=text)
        if bold:
            cell.font = Font(bold=True)

    # Row 3: empty (visual separator)
    # Row 4: Headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    fills = {
        'valid': PatternFill('solid', fgColor='ECFDF5'),
        'supplementary': PatternFill('solid', fgColor='DBEAFE'),
        'not_attended': PatternFill('solid', fgColor='FEF2F2'),
        'not_registered': PatternFill('solid', fgColor='FFF7ED'),
        'no_profile': PatternFill('solid', fgColor='FEF3C7'),
    }

    for idx, r in enumerate(rows, start=1):
        excel_row = idx + 4
        profile = r.get('profile')
        unit = (profile.unit if profile else '') or ''
        dept = (profile.department if profile else '') or ''
        # scan_time lưu UTC (USE_TZ=True) — convert sang giờ VN trước khi format.
        scan_str = timezone.localtime(r['scan_time']).strftime('%H:%M:%S') if r['scan_time'] else '—'
        values = [
            idx,
            r['display_name'],
            r['employee_code'],
            r.get('quantity', 0),
            unit,
            dept,
            scan_str,
            r['status_label'],
            r['type'],
        ]
        fill = fills.get(r['status'])
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=v)
            if fill:
                cell.fill = fill

    widths = [6, 28, 14, 12, 24, 24, 16, 18, 14]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord('A') + col_idx - 1)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------- Gửi qua NetChat DM ----------

def _get_netchat_config():
    url_cfg = SystemConfig.objects.filter(key='netchat_url').first()
    token_cfg = SystemConfig.objects.filter(key='netchat_token').first()
    if not url_cfg or not token_cfg:
        return None
    return {
        'url': url_cfg.value.strip().rstrip('/'),
        'token': token_cfg.value.strip(),
    }


def send_excel_to_recipients(target_date, file_bytes, recipient_codes, message=None):
    """Gửi file Excel qua DM NetChat cho danh sách MNV.

    Trả về dict {success: [...], failed: [(code, reason), ...]}.
    """
    if message is None:
        message = build_report_caption(target_date)
    cfg = _get_netchat_config()
    if not cfg:
        return {
            'success': [],
            'failed': [(c, 'Bot chưa được cấu hình URL/Token (vào Profile admin để thiết lập)') for c in recipient_codes],
        }

    headers = {
        'Authorization': f"Bearer {cfg['token']}",
        'User-Agent': 'curl/8.7.1',
    }

    # 1. Lấy bot_id
    try:
        r_me = requests.get(f"{cfg['url']}/api/v4/users/me", headers=headers, timeout=10)
        if r_me.status_code != 200:
            return {'success': [], 'failed': [(c, f'Bot không đăng nhập được: HTTP {r_me.status_code}') for c in recipient_codes]}
        bot_id = r_me.json().get('id')
    except requests.RequestException as e:
        return {'success': [], 'failed': [(c, f'Lỗi kết nối NetChat: {e}') for c in recipient_codes]}

    profiles = {
        (p.employee_code or '').strip(): p
        for p in UserProfile.objects.filter(employee_code__in=recipient_codes)
    }

    filename = f'tham_gia_{target_date.strftime("%Y-%m-%d")}.xlsx'
    success = []
    failed = []

    for code in recipient_codes:
        profile = profiles.get(code)
        if not profile or not profile.email:
            failed.append((code, 'Không tìm thấy tài khoản hoặc chưa có email'))
            continue

        username = profile.email.split('@')[0].strip().lower()
        try:
            r_user = requests.get(
                f"{cfg['url']}/api/v4/users/username/{username}",
                headers=headers, timeout=10
            )
            if r_user.status_code != 200:
                failed.append((code, 'Không tìm thấy user trên NetChat'))
                continue
            user_mm_id = r_user.json().get('id')

            r_chan = requests.post(
                f"{cfg['url']}/api/v4/channels/direct",
                headers={**headers, 'Content-Type': 'application/json'},
                json=[bot_id, user_mm_id], timeout=10
            )
            if r_chan.status_code not in (200, 201):
                failed.append((code, f'Không mở được kênh DM (HTTP {r_chan.status_code})'))
                continue
            channel_id = r_chan.json().get('id')

            # Upload file (multipart) — KHÔNG set Content-Type=json.
            files_payload = {
                'files': (filename, file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            }
            r_file = requests.post(
                f"{cfg['url']}/api/v4/files",
                headers=headers,
                data={'channel_id': channel_id},
                files=files_payload,
                timeout=30,
            )
            if r_file.status_code not in (200, 201):
                failed.append((code, f'Upload file thất bại (HTTP {r_file.status_code})'))
                continue
            file_infos = r_file.json().get('file_infos') or []
            if not file_infos:
                failed.append((code, 'Mattermost không trả file_id'))
                continue
            file_id = file_infos[0].get('id')

            # Tạo post kèm file.
            r_post = requests.post(
                f"{cfg['url']}/api/v4/posts",
                headers={**headers, 'Content-Type': 'application/json'},
                json={'channel_id': channel_id, 'message': message, 'file_ids': [file_id]},
                timeout=10,
            )
            if r_post.status_code not in (200, 201):
                failed.append((code, f'Gửi post thất bại (HTTP {r_post.status_code})'))
                continue

            success.append(code)
        except requests.RequestException as e:
            failed.append((code, f'Lỗi mạng: {e}'))

    return {'success': success, 'failed': failed}


def send_excel_to_channel(target_date, file_bytes, channel_id, message=None):
    """Đăng file Excel báo cáo trực tiếp vào 1 channel NetChat theo channel_id.

    Bot phải là thành viên của channel đó. Trả về {'ok': bool, 'message': str}.
    """
    if message is None:
        message = build_report_caption(target_date)
    cfg = _get_netchat_config()
    if not cfg:
        return {'ok': False, 'message': 'Bot chưa được cấu hình URL/Token (vào Profile admin để thiết lập).'}

    channel_id = (channel_id or '').strip()
    if not channel_id:
        return {'ok': False, 'message': 'Chưa cấu hình Channel ID.'}

    headers = {
        'Authorization': f"Bearer {cfg['token']}",
        'User-Agent': 'curl/8.7.1',
    }
    filename = f'tham_gia_{target_date.strftime("%Y-%m-%d")}.xlsx'

    try:
        # 1. Upload file vào channel.
        r_file = requests.post(
            f"{cfg['url']}/api/v4/files",
            headers=headers,
            data={'channel_id': channel_id},
            files={'files': (
                filename, file_bytes,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )},
            timeout=30,
        )
        if r_file.status_code not in (200, 201):
            return {'ok': False, 'message': (
                f'Upload file thất bại (HTTP {r_file.status_code}). '
                'Kiểm tra Channel ID đúng chưa và Bot đã là thành viên channel chưa.'
            )}
        file_infos = r_file.json().get('file_infos') or []
        if not file_infos:
            return {'ok': False, 'message': 'NetChat không trả về file_id.'}
        file_id = file_infos[0].get('id')

        # 2. Tạo post kèm file trong channel.
        r_post = requests.post(
            f"{cfg['url']}/api/v4/posts",
            headers={**headers, 'Content-Type': 'application/json'},
            json={'channel_id': channel_id, 'message': message, 'file_ids': [file_id]},
            timeout=10,
        )
        if r_post.status_code not in (200, 201):
            return {'ok': False, 'message': (
                f'Gửi post thất bại (HTTP {r_post.status_code}). '
                'Bot có thể chưa là thành viên của channel.'
            )}
        return {'ok': True, 'message': f'Đã đăng báo cáo vào channel ({channel_id}).'}
    except requests.RequestException as e:
        return {'ok': False, 'message': f'Lỗi kết nối NetChat: {e}'}


def send_participation_excel(target_date, file_bytes, rows=None):
    """Gửi báo cáo Tham gia qua NetChat theo hình thức đã cấu hình (dm | channel).

    Dùng chung cho view gửi tay và management command auto-send. Nếu truyền
    `rows` thì caption tin nhắn kèm số liệu tổng hợp.
    Trả về {'ok': bool, 'message': str}.
    """
    caption = build_report_caption(target_date, rows)
    mode = get_send_mode()

    if mode == 'channel':
        channel_id = get_channel_id()
        if not channel_id:
            return {'ok': False, 'message': 'Chưa cấu hình Channel ID. Vào Cài đặt để nhập.'}
        return send_excel_to_channel(target_date, file_bytes, channel_id, caption)

    # mode == 'dm'
    recipients = get_recipients()
    if not recipients:
        return {'ok': False, 'message': 'Chưa cấu hình người nhận. Vào Cài đặt để thêm mã NV.'}

    result = send_excel_to_recipients(target_date, file_bytes, recipients, caption)
    sent = len(result['success'])
    total = len(recipients)
    if sent == 0:
        message = f'Gửi thất bại cho cả {total} người.'
    else:
        message = f'Đã gửi cho {sent}/{total} người.'
    if result['failed']:
        fails = '; '.join(f'{c}: {r}' for c, r in result['failed'])
        message += f'\nLỗi: {fails}'
    return {'ok': sent > 0, 'message': message}
